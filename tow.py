import pymysql, xlwt


def export_excel(hero):
    # 连接数据库，查询数据
    host, user, passwd, db = '127.0.0.1', 'root', '12345678', 'RUNOOB'
    conn = pymysql.connect(user=user, host=host, port=3306, passwd=passwd, db=db, charset='utf8')
    cur = conn.cursor()
    sql = 'select * from %s' % hero
    cur.execute(sql)  # 返回受影响的行数

    fields = [field[0] for field in cur.description]  # 获取所有字段名
    all_data = cur.fetchall()  # 所有数据

    # 写入excel
    book = xlwt.Workbook()
    sheet = book.add_sheet('sheet1')

    for col, field in enumerate(fields):
        sheet.write(0, col, field)

    row = 1
    for data in all_data:
        for col, field in enumerate(data):
            sheet.write(row, col, field)
        row += 1
    book.save("%s.xls" % hero)


if __name__ == '__main__':
    export_excel('hero')
    
#文件较多的方式导出法
# coding=utf-8
import pymysql
import openpyxl


def export_excel(hero):
    # 连接数据库，查询数据
    host, user, passwd, db = '127.0.0.1', 'root', '12345678', 'RUNOOB'
    conn = pymysql.connect(user=user, host=host, port=3306, passwd=passwd, db=db, charset='utf8')
    cur = conn.cursor()
    sql = 'select * from %s' % hero
    cur.execute(sql)  # 返回受影响的行数

    fields = [field[0] for field in cur.description]  # 获取所有字段名
    all_data = cur.fetchall()  # 所有数据

    # 写入excel
    book = openpyxl.Workbook()
    sheet = book.create_sheet(index=1)
    ws = book.active

    # for col, field in enumerate(fields):
    #     sheet.write(0, col, field)

    i = 1
    r = 1
    for line in all_data:
        for col in range(1, len(line) + 1):
            ColNum = r
            ws.cell(row=r, column=col).value = line[col - 1]
        i += 1
        r += 1
    # 工作簿保存到磁盘
    book.save('%s.xlsx' % hero)


if __name__ == '__main__':
    export_excel('hero')
